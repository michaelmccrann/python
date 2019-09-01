#!/usr/bin/env python3

from openpyxl import Workbook
from openpyxl import load_workbook
import utilities

wb = load_workbook(filename = 'data/test_spreadshhet.xlsx')
ws = wb.active

for row in ws.iter_rows():
      test_data=utilities.get_test_data(row)
      print (test_data.Total)    

